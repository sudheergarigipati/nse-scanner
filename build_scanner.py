#!/usr/bin/env python3
# ═══════════════════════════════════════════════════════════════
#  build_scanner.py
#  Reads NSE 5yr Excel → finds highest volume day per stock
#  → generates NSE_Volume_Scanner.html
#  Run manually: python3 build_scanner.py
#  Run weekly:   cron will call this automatically
# ═══════════════════════════════════════════════════════════════

import openpyxl
import json
import os
import glob
import sys
from datetime import datetime

BASE_DIR    = os.path.expanduser('~/nse-scanner')
OUTPUT_HTML = os.path.join(BASE_DIR, 'NSE_Volume_Scanner.html')
SERVE_DIR   = os.path.join(BASE_DIR, 'www')

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def find_excel():
    patterns = [
        os.path.join(BASE_DIR, 'NSE_5yr*.xlsx'),
        os.path.join(BASE_DIR, '*.xlsx'),
    ]
    files = []
    for p in patterns:
        files.extend(glob.glob(p))
    if not files:
        raise FileNotFoundError(
            f"No Excel file found in {BASE_DIR}\n"
            "Upload your NSE_5yr_Historical_Data_FULL.xlsx file first."
        )
    latest = max(files, key=os.path.getmtime)
    log(f"Found Excel: {os.path.basename(latest)}")
    return latest

def build_raw_data(excel_path):
    log("Reading SUMMARY sheet...")
    wb = openpyxl.load_workbook(excel_path, read_only=True)

    # Find summary sheet (handles emoji in sheet name)
    summary_sheet = None
    for name in wb.sheetnames:
        if 'SUMMARY' in name.upper():
            summary_sheet = name
            break
    if not summary_sheet:
        raise ValueError(f"No SUMMARY sheet found. Sheets: {wb.sheetnames}")

    ws   = wb[summary_sheet]
    rows = list(ws.iter_rows(values_only=True))
    log(f"Total rows in SUMMARY: {len(rows)}")

    raw = []
    skipped = 0
    for row in rows[1:]:  # skip header
        try:
            sym, fd, td, days, mv, mvd, hl, hh, hc, lc, ld = row
            if not sym or not mv or not hl or not hh or not lc:
                skipped += 1
                continue
            raw.append({
                's':   str(sym).strip(),
                'mvd': str(mvd).strip(),
                'mv':  int(mv),
                'hl':  float(hl),
                'hh':  float(hh),
                'hc':  float(hc),
                'lc':  float(lc),
                'ld':  str(ld).strip(),
                'td':  int(days) if days else 0
            })
        except Exception as e:
            skipped += 1
            continue

    log(f"Loaded {len(raw)} stocks (skipped {skipped} incomplete rows)")
    return raw

def generate_html(raw):
    log("Generating HTML scanner...")
    raw_js = json.dumps(raw, separators=(',', ':'))
    now_str = datetime.now().strftime('%d %b %Y %H:%M')

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>NSE Volume Scanner Ver1.0</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:#0a0f1a;color:#e2e8f0;min-height:100vh}}
header{{background:#0f172a;border-bottom:1px solid #1e293b;padding:14px 24px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:200}}
.logo{{font-size:18px;font-weight:700;color:#38bdf8}}.logo span{{color:#f59e0b}}
.ver{{font-size:11px;color:#475569;margin-left:6px;vertical-align:middle}}
#statusBar{{font-size:12px;color:#64748b}}
.updated{{font-size:11px;color:#334155;margin-left:12px}}
.tabs{{display:flex;background:#0f172a;border-bottom:2px solid #1e293b;padding:0 24px;position:sticky;top:57px;z-index:150}}
.tab{{padding:12px 28px;font-size:13px;font-weight:600;cursor:pointer;border-bottom:3px solid transparent;margin-bottom:-2px;color:#64748b;display:flex;align-items:center;gap:8px}}
.tab.bull-tab.active{{color:#22c55e;border-bottom-color:#22c55e}}
.tab.bear-tab.active{{color:#ef4444;border-bottom-color:#ef4444}}
.tab-count{{font-size:11px;padding:2px 8px;border-radius:20px}}
.bull-tab .tab-count{{background:#052e16;color:#22c55e}}
.bear-tab .tab-count{{background:#1a0808;color:#ef4444}}
.controls{{background:#0f172a;border-bottom:1px solid #1e293b;padding:12px 24px;display:flex;gap:12px;align-items:center;flex-wrap:wrap}}
.ctrl-group{{display:flex;align-items:center;gap:8px}}
label{{font-size:12px;color:#94a3b8}}
select,input[type=number]{{background:#1e293b;border:1px solid #334155;color:#e2e8f0;border-radius:6px;padding:6px 10px;font-size:13px;outline:none}}
.btn{{padding:8px 18px;border-radius:6px;border:none;font-size:13px;font-weight:600;cursor:pointer}}
.btn-bull{{background:#16a34a;color:#fff}}.btn-bull:hover{{background:#22c55e}}
.btn-bear{{background:#dc2626;color:#fff}}.btn-bear:hover{{background:#ef4444}}
.btn-both{{background:#0ea5e9;color:#fff}}.btn-both:hover{{background:#38bdf8}}
.btn:disabled{{background:#334155!important;color:#64748b;cursor:not-allowed}}
.btn-refresh{{background:#0f172a;border:1px solid #334155;color:#94a3b8;font-size:12px;padding:6px 14px}}
.btn-secondary{{background:#1e293b;color:#94a3b8;border:1px solid #334155}}
.stats-row{{display:grid;gap:12px;padding:14px 24px;grid-template-columns:1fr 1fr 1fr 1fr}}
.stat-card{{background:#0f172a;border:1px solid #1e293b;border-radius:8px;padding:11px 14px}}
.stat-val{{font-size:24px;font-weight:700;color:#f1f5f9}}
.stat-lbl{{font-size:10px;color:#64748b;margin-top:2px;text-transform:uppercase}}
.s-blue .stat-val{{color:#38bdf8}}.s-green .stat-val{{color:#22c55e}}
.s-amber .stat-val{{color:#f59e0b}}.s-red .stat-val{{color:#ef4444}}
.s-purple .stat-val{{color:#a78bfa}}
.progress-wrap{{padding:0 24px 10px}}
.progress-bar{{height:5px;background:#1e293b;border-radius:3px;overflow:hidden}}
.progress-fill-bull{{height:100%;background:linear-gradient(90deg,#16a34a,#22c55e);border-radius:3px;transition:width .15s;width:0%}}
.progress-fill-bear{{height:100%;background:linear-gradient(90deg,#dc2626,#ef4444);border-radius:3px;transition:width .15s;width:0%}}
.progress-label{{font-size:11px;color:#64748b;margin-top:5px;display:flex;justify-content:space-between}}
.info-box{{border-radius:8px;padding:9px 16px;margin:8px 24px;font-size:12px;line-height:1.6;display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap}}
.info-blue{{background:#0c1a2e;border:1px solid #0ea5e9;color:#7dd3fc}}
.info-success{{background:#052e16;border:1px solid #166534;color:#86efac}}
.info-bear-success{{background:#1a0808;border:1px solid #991b1b;color:#fca5a5}}
.filter-row{{display:flex;gap:8px;padding:0 24px 10px;flex-wrap:wrap;align-items:center}}
.chip{{padding:4px 14px;border-radius:20px;border:1px solid #334155;background:transparent;color:#94a3b8;font-size:12px;cursor:pointer}}
.chip-bull.active{{background:#16a34a;border-color:#16a34a;color:#fff}}
.chip-bear.active{{background:#dc2626;border-color:#dc2626;color:#fff}}
.results{{padding:0 24px 32px}}
.section-title{{font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.08em;margin:14px 0 8px;display:flex;align-items:center;gap:8px}}
.badge{{background:#1e293b;color:#94a3b8;border-radius:20px;padding:2px 8px;font-size:11px}}
table{{width:100%;border-collapse:collapse;font-size:13px;margin-bottom:6px}}
th{{background:#0d1424;color:#64748b;font-weight:600;font-size:11px;text-transform:uppercase;padding:9px 11px;text-align:left;border-bottom:1px solid #1e293b;position:sticky;top:110px;z-index:5}}
td{{padding:8px 11px;border-bottom:1px solid #0f172a}}
tr:hover td{{background:#0f172a99}}
.sym{{font-weight:700;color:#f1f5f9;font-size:13px}}
.c-green{{color:#22c55e}}.c-red{{color:#ef4444}}.c-amber{{color:#f59e0b}}
.c-blue{{color:#38bdf8}}.c-muted{{color:#64748b}}.c-purple{{color:#a78bfa}}
.tag{{display:inline-block;padding:2px 7px;border-radius:4px;font-size:10px;font-weight:700;margin-top:2px}}
.tag-short{{background:#1a0808;color:#ef4444;border:1px solid #991b1b}}
.tag-watch-bear{{background:#1c0a00;color:#fb923c;border:1px solid #9a3412}}
.tag-weak{{background:#0f0f1a;color:#a78bfa;border:1px solid #6d28d9}}
.tag-buy{{background:#052e16;color:#22c55e;border:1px solid #166534}}
.tag-watch{{background:#1c1917;color:#f59e0b;border:1px solid #92400e}}
.tag-broken{{background:#1a0808;color:#ef4444;border:1px solid #991b1b}}
.tv-link{{color:#475569;text-decoration:none;font-size:11px;padding:2px 7px;border:1px solid #1e293b;border-radius:4px}}
.tv-link:hover{{color:#38bdf8;border-color:#38bdf8}}
.empty{{text-align:center;padding:48px 24px;color:#475569}}
.price-src{{font-size:10px;color:#475569}}
.live-dot{{display:inline-block;width:6px;height:6px;border-radius:50%;background:#22c55e;margin-right:3px;vertical-align:middle}}
.cached-dot{{display:inline-block;width:6px;height:6px;border-radius:50%;background:#f59e0b;margin-right:3px;vertical-align:middle}}
.panel{{display:none}}.panel.active{{display:block}}
</style>
</head>
<body>

<header>
  <div>
    <span class="logo">NSE <span>Volume</span> Scanner <span class="ver">Ver1.0</span></span>
    <span class="updated">Data updated: {now_str}</span>
  </div>
  <div id="statusBar">{len(raw)} stocks loaded</div>
</header>

<div class="tabs">
  <div class="tab bull-tab active" id="tabBull" onclick="switchTab('bull')">
    📈 Bullish Scanner <span class="tab-count" id="bullTabCount">—</span>
  </div>
  <div class="tab bear-tab" id="tabBear" onclick="switchTab('bear')">
    📉 Bearish Scanner <span class="tab-count" id="bearTabCount">—</span>
  </div>
</div>

<div class="controls">
  <div class="ctrl-group">
    <label>HV within</label>
    <select id="daysFilter">
      <option value="30">30 days</option>
      <option value="60">60 days</option>
      <option value="90">90 days</option>
      <option value="180" selected>180 days</option>
      <option value="365">1 year</option>
      <option value="9999">All time</option>
    </select>
  </div>
  <div class="ctrl-group" id="bullBufWrap">
    <label>Max % above support</label>
    <input type="number" id="bullBuf" value="15" min="1" max="50" style="width:65px">
  </div>
  <div class="ctrl-group" id="bearBufWrap" style="display:none">
    <label>Max % below resistance</label>
    <input type="number" id="bearBuf" value="15" min="1" max="50" style="width:65px">
  </div>
  <button class="btn btn-both" id="scanBtn" onclick="startScan()">&#9654;&nbsp; Run Scan</button>
  <button class="btn btn-refresh" id="refreshBtn" onclick="refreshLive()" style="display:none">⟳ Refresh Live</button>
  <button class="btn btn-secondary" onclick="exportCSV()">&#8595; Export CSV</button>
</div>

<div class="panel active" id="panelBull">
  <div class="stats-row">
    <div class="stat-card s-blue"><div class="stat-val" id="b-total">{len(raw)}</div><div class="stat-lbl">Total Stocks</div></div>
    <div class="stat-card s-green"><div class="stat-val" id="b-buy">—</div><div class="stat-lbl">▲ Buy Setups</div></div>
    <div class="stat-card s-amber"><div class="stat-val" id="b-watch">—</div><div class="stat-lbl">Watchlist</div></div>
    <div class="stat-card s-red"><div class="stat-val" id="b-broken">—</div><div class="stat-lbl">Support Broken</div></div>
  </div>
  <div class="progress-wrap" id="b-prog" style="display:none">
    <div class="progress-bar"><div class="progress-fill-bull" id="b-progFill"></div></div>
    <div class="progress-label"><span id="b-progLbl">Starting...</span><span id="b-progEta"></span></div>
  </div>
  <div id="b-info"></div>
  <div class="filter-row" id="b-filterRow" style="display:none">
    <span style="font-size:12px;color:#64748b">Show:</span>
    <button class="chip chip-bull active" onclick="setView('bull',this,'all')">All</button>
    <button class="chip chip-bull" onclick="setView('bull',this,'buy')">▲ Buy Only</button>
    <button class="chip chip-bull" onclick="setView('bull',this,'watch')">Watch</button>
    <button class="chip chip-bull" onclick="setView('bull',this,'broken')">Broken</button>
  </div>
  <div class="results" id="b-results">
    <div class="empty">
      <div style="font-size:40px;margin-bottom:12px">📈</div>
      <div style="font-size:16px;font-weight:600;color:#64748b;margin-bottom:6px">Bullish Volume Scanner</div>
      <div style="font-size:13px">Finds stocks near 5-year highest-volume day LOW (support holding)</div>
    </div>
  </div>
</div>

<div class="panel" id="panelBear">
  <div class="stats-row">
    <div class="stat-card s-blue"><div class="stat-val" id="r-total">{len(raw)}</div><div class="stat-lbl">Total Stocks</div></div>
    <div class="stat-card s-red"><div class="stat-val" id="r-short">—</div><div class="stat-lbl">▼ Short Setups</div></div>
    <div class="stat-card s-amber"><div class="stat-val" id="r-watch">—</div><div class="stat-lbl">Watch</div></div>
    <div class="stat-card s-purple"><div class="stat-val" id="r-weak">—</div><div class="stat-lbl">Weak</div></div>
  </div>
  <div class="progress-wrap" id="r-prog" style="display:none">
    <div class="progress-bar"><div class="progress-fill-bear" id="r-progFill"></div></div>
    <div class="progress-label"><span id="r-progLbl">Starting...</span><span id="r-progEta"></span></div>
  </div>
  <div id="r-info"></div>
  <div class="filter-row" id="r-filterRow" style="display:none">
    <span style="font-size:12px;color:#64748b">Show:</span>
    <button class="chip chip-bear active" onclick="setView('bear',this,'all')">All</button>
    <button class="chip chip-bear" onclick="setView('bear',this,'short')">▼ Short Only</button>
    <button class="chip chip-bear" onclick="setView('bear',this,'watch')">Watch</button>
    <button class="chip chip-bear" onclick="setView('bear',this,'weak')">Weak</button>
  </div>
  <div class="results" id="r-results">
    <div class="empty">
      <div style="font-size:40px;margin-bottom:12px">📉</div>
      <div style="font-size:16px;font-weight:600;color:#64748b;margin-bottom:6px">Bearish Volume Scanner</div>
      <div style="font-size:13px">Finds stocks near 5-year highest-volume day HIGH (resistance rejecting)</div>
    </div>
  </div>
</div>

<script>
const RAW={raw_js};
let bullResults=[],bearResults=[];
let bullView='all',bearView='all',activeTab='bull';
let isRefreshing=false;
let scanParams={{days:180,bullBuf:15,bearBuf:15}};

function switchTab(tab){{
  activeTab=tab;
  document.getElementById('tabBull').classList.toggle('active',tab==='bull');
  document.getElementById('tabBear').classList.toggle('active',tab==='bear');
  document.getElementById('panelBull').classList.toggle('active',tab==='bull');
  document.getElementById('panelBear').classList.toggle('active',tab==='bear');
  document.getElementById('bullBufWrap').style.display=tab==='bull'?'flex':'none';
  document.getElementById('bearBufWrap').style.display=tab==='bear'?'flex':'none';
  document.getElementById('scanBtn').className='btn '+(tab==='bull'?'btn-bull':'btn-bear');
}}

function scoreBull(s,price,daysSince){{
  let sc=0;
  const pa=(price-s.hl)/s.hl*100,hr=s.hh-s.hl,cp=hr>0?(s.hc-s.hl)/hr:0.5;
  if(cp>=0.5)sc+=15;if(price>=s.hl)sc+=20;
  if(pa<=3)sc+=20;else if(pa<=7)sc+=12;else if(pa<=12)sc+=6;
  if(daysSince<=7)sc+=20;else if(daysSince<=30)sc+=14;else if(daysSince<=90)sc+=8;else sc+=3;
  if(s.mv>100000000)sc+=10;else if(s.mv>50000000)sc+=7;else if(s.mv>10000000)sc+=4;
  const rng=hr/s.hl*100;if(rng>10)sc+=10;else if(rng>5)sc+=5;
  return Math.min(sc,100);
}}

function scoreBear(s,price,daysSince){{
  let sc=0;
  const pb=(s.hh-price)/s.hh*100,hr=s.hh-s.hl,cp=hr>0?(s.hc-s.hl)/hr:0.5;
  if(cp<=0.5)sc+=15;if(price<s.hh)sc+=20;
  if(pb<=3)sc+=20;else if(pb<=7)sc+=14;else if(pb<=12)sc+=8;else if(pb<=20)sc+=3;
  if(daysSince<=7)sc+=20;else if(daysSince<=30)sc+=14;else if(daysSince<=90)sc+=8;else sc+=3;
  if(s.mv>100000000)sc+=10;else if(s.mv>50000000)sc+=7;else if(s.mv>10000000)sc+=4;
  const rng=hr/s.hl*100;if(rng>10)sc+=10;else if(rng>5)sc+=5;
  return Math.min(sc,100);
}}

function buildBull(s,price,date,daysSince,isLive,volume=0){{
  const pa=(price-s.hl)/s.hl*100,hr=s.hh-s.hl,cp=hr>0?(s.hc-s.hl)/hr:0.5;
  const intact=price>=s.hl*0.99,upside=((s.hh-price)/price*100).toFixed(1);
  const score=scoreBull(s,price,daysSince);
  const grade=score>=65&&intact?'buy':intact?'watch':'broken';
  return {{symbol:s.s,hvDate:s.mvd,daysSince,hvLow:s.hl.toFixed(2),hvHigh:s.hh.toFixed(2),
    hvBullish:cp>=0.5,currentPrice:price.toFixed(2),priceDate:date,
    pctAbove:pa.toFixed(2),upside,stopLoss:(s.hl*0.98).toFixed(2),
    score,grade,intact,isLive,volume}};
}}

function buildBear(s,price,date,daysSince,isLive,volume=0){{
  const pb=(s.hh-price)/s.hh*100,hr=s.hh-s.hl,cp=hr>0?(s.hc-s.hl)/hr:0.5;
  const downside=((price-s.hl)/price*100).toFixed(1);
  const score=scoreBear(s,price,daysSince);
  const grade=score>=65?'short':pb<=15?'watch':'weak';
  return {{symbol:s.s,hvDate:s.mvd,daysSince,hvLow:s.hl.toFixed(2),hvHigh:s.hh.toFixed(2),
    hvBearish:cp<=0.5,currentPrice:price.toFixed(2),priceDate:date,
    pctBelow:pb.toFixed(2),downside,stopLoss:(s.hh*1.02).toFixed(2),
    score,grade,isLive,volume}};
}}

function runStoredBull(days,buf){{
  const now=Date.now();bullResults=[];
  const cands=RAW.filter(s=>Math.round((now-new Date(s.mvd).getTime())/86400000)<=days);
  for(const s of cands){{
    if(!s.lc||s.lc<=0)continue;
    const d=Math.round((now-new Date(s.mvd).getTime())/86400000);
    const pa=(s.lc-s.hl)/s.hl*100;
    if(pa<-1||pa>buf)continue;
    bullResults.push(buildBull(s,s.lc,s.ld,d,false));
  }}
  bullResults.sort((a,b)=>b.score-a.score);
  return cands.length;
}}

function runStoredBear(days,buf){{
  const now=Date.now();bearResults=[];
  const cands=RAW.filter(s=>Math.round((now-new Date(s.mvd).getTime())/86400000)<=days);
  for(const s of cands){{
    if(!s.lc||s.lc<=0)continue;
    const d=Math.round((now-new Date(s.mvd).getTime())/86400000);
    const pb=(s.hh-s.lc)/s.hh*100;
    if(s.lc>=s.hh||pb>buf)continue;
    bearResults.push(buildBear(s,s.lc,s.ld,d,false));
  }}
  bearResults.sort((a,b)=>b.score-a.score);
  return cands.length;
}}

async function fetchLive(sym){{
  const tryOne=async url=>{{
    const r=await fetch(url,{{signal:AbortSignal.timeout(4000)}});
    if(!r.ok)throw 0;
    const d=await r.json();
    const q=d?.quoteResponse?.result?.[0];
    if(!q||!q.regularMarketPrice)throw 0;
    return {{price:q.regularMarketPrice,volume:q.regularMarketVolume||0,
             date:new Date(q.regularMarketTime*1000).toLocaleDateString('en-IN')}};
  }};
  return Promise.any([
    tryOne(`https://query2.finance.yahoo.com/v7/finance/quote?symbols=${{sym}}.NS`),
    tryOne(`https://query1.finance.yahoo.com/v7/finance/quote?symbols=${{sym}}.NS`),
  ]).catch(()=>null);
}}

async function refreshLive(){{
  if(isRefreshing)return;
  isRefreshing=true;
  const isBull=activeTab==='bull';
  const results=isBull?bullResults:bearResults;
  const symMap=Object.fromEntries(RAW.map(s=>[s.s,s]));
  const now=Date.now(),buf=isBull?scanParams.bullBuf:scanParams.bearBuf;
  const btn=document.getElementById('refreshBtn');
  btn.disabled=true;btn.textContent='⟳ Fetching...';
  const pfx=isBull?'b':'r';
  document.getElementById(pfx+'-prog').style.display='block';
  document.getElementById(pfx+'-progFill').style.width='0%';
  const syms=results.map(r=>r.symbol);
  const BATCH=20;let done=0,liveHits=0;
  const newRes=[],resMap=new Map(results.map(r=>[r.symbol,r]));
  const start=Date.now();
  for(let i=0;i<syms.length;i+=BATCH){{
    const batch=syms.slice(i,Math.min(i+BATCH,syms.length));
    const fetched=await Promise.allSettled(batch.map(sym=>fetchLive(sym)));
    fetched.forEach((res,j)=>{{
      const sym=batch[j];done++;
      const pct=Math.round(done/syms.length*100);
      const eta=done>2?Math.round((Date.now()-start)/1000/done*(syms.length-done)):0;
      document.getElementById(pfx+'-progFill').style.width=pct+'%';
      document.getElementById(pfx+'-progLbl').textContent=`Live prices ${{done}}/${{syms.length}} (${{pct}}%)`;
      document.getElementById(pfx+'-progEta').textContent=eta>0?`~${{eta}}s left`:'';
      if(res.status==='fulfilled'&&res.value){{
        const lv=res.value,s=symMap[sym];if(!s)return;
        const d=Math.round((now-new Date(s.mvd).getTime())/86400000);
        if(isBull){{const pa=(lv.price-s.hl)/s.hl*100;if(pa>=-1&&pa<=buf){{newRes.push(buildBull(s,lv.price,lv.date,d,true,lv.volume));liveHits++;}}}}
        else{{const pb=(s.hh-lv.price)/s.hh*100;if(lv.price<s.hh&&pb<=buf){{newRes.push(buildBear(s,lv.price,lv.date,d,true,lv.volume));liveHits++;}}}}
      }}else{{const ex=resMap.get(sym);if(ex)newRes.push(ex);}}
    }});
  }}
  newRes.sort((a,b)=>b.score-a.score);
  if(isBull){{bullResults=newRes;updateBullStats();}}else{{bearResults=newRes;updateBearStats();}}
  document.getElementById(pfx+'-prog').style.display='none';
  document.getElementById(pfx+'-info').innerHTML=
    `<div class="info-box ${{isBull?'info-success':'info-bear-success'}}">
      ✅ Live prices updated · ${{liveHits}}/${{syms.length}} live · ${{syms.length-liveHits}} stored fallback
    </div>`;
  document.getElementById('statusBar').textContent=`Live refresh done — ${{newRes.length}} ${{isBull?'bullish':'bearish'}} stocks`;
  btn.disabled=false;btn.textContent='⟳ Refresh Live';isRefreshing=false;
  renderTable(isBull?'bull':'bear');
}}

function updateBullStats(){{
  document.getElementById('b-buy').textContent=bullResults.filter(r=>r.grade==='buy').length;
  document.getElementById('b-watch').textContent=bullResults.filter(r=>r.grade==='watch').length;
  document.getElementById('b-broken').textContent=bullResults.filter(r=>r.grade==='broken').length;
  document.getElementById('bullTabCount').textContent=bullResults.length;
}}
function updateBearStats(){{
  document.getElementById('r-short').textContent=bearResults.filter(r=>r.grade==='short').length;
  document.getElementById('r-watch').textContent=bearResults.filter(r=>r.grade==='watch').length;
  document.getElementById('r-weak').textContent=bearResults.filter(r=>r.grade==='weak').length;
  document.getElementById('bearTabCount').textContent=bearResults.length;
}}

async function startScan(){{
  const btn=document.getElementById('scanBtn');
  btn.disabled=true;btn.textContent='Scanning...';
  const isBull=activeTab==='bull',pfx=isBull?'b':'r';
  scanParams.days=parseInt(document.getElementById('daysFilter').value);
  scanParams.bullBuf=parseFloat(document.getElementById('bullBuf').value);
  scanParams.bearBuf=parseFloat(document.getElementById('bearBuf').value);
  document.getElementById(pfx+'-filterRow').style.display='none';
  document.getElementById('refreshBtn').style.display='none';
  document.getElementById(pfx+'-info').innerHTML='';
  document.getElementById(pfx+'-results').innerHTML='';
  const t0=performance.now();
  let total=isBull?runStoredBull(scanParams.days,scanParams.bullBuf):runStoredBear(scanParams.days,scanParams.bearBuf);
  const t1=performance.now();
  if(isBull)updateBullStats();else updateBearStats();
  const matched=isBull?bullResults.length:bearResults.length;
  document.getElementById(pfx+'-filterRow').style.display='flex';
  document.getElementById('refreshBtn').style.display='inline-block';
  document.getElementById('refreshBtn').disabled=false;
  document.getElementById(pfx+'-info').innerHTML=
    `<div class="info-box info-blue">📦 Stored prices · ${{matched}} stocks found in ${{Math.round(t1-t0)}}ms · Click Refresh Live to update</div>`;
  document.getElementById('statusBar').textContent=`Stored scan · ${{total}} checked · ${{matched}} matched`;
  btn.disabled=false;btn.textContent='↺ Re-scan';
  renderTable(isBull?'bull':'bear');
  await refreshLive();
}}

function setView(mode,el,v){{
  if(mode==='bull')bullView=v;else bearView=v;
  el.closest('.filter-row').querySelectorAll('.chip').forEach(c=>c.classList.remove('active'));
  el.classList.add('active');renderTable(mode);
}}

function mkBullRows(arr){{return arr.map(r=>`
  <tr>
    <td><div class="sym">${{r.symbol}}</div><span class="tag tag-${{r.grade}}">${{r.grade==='buy'?'BUY':r.grade==='watch'?'WATCH':'BROKEN'}}</span></td>
    <td><span style="font-weight:700;font-size:15px;color:${{r.score>=65?'#22c55e':r.score>=45?'#f59e0b':'#64748b'}}">${{r.score}}</span><span class="c-muted">/100</span></td>
    <td class="c-muted" style="font-size:11px">${{r.hvDate}}</td>
    <td class="${{r.daysSince<=7?'c-green':r.daysSince<=30?'c-amber':'c-muted'}}">${{r.daysSince}}d</td>
    <td style="font-weight:600;color:#f1f5f9">₹${{r.hvLow}}</td>
    <td class="c-muted">₹${{r.hvHigh}}</td>
    <td><div style="font-weight:700;font-size:14px;color:#38bdf8">₹${{r.currentPrice}}</div>
        <div class="price-src">${{r.isLive?'<span class="live-dot"></span>live':'<span class="cached-dot"></span>'+r.priceDate}}${{r.volume>0?' · Vol:'+r.volume.toLocaleString():''}}</div></td>
    <td class="${{parseFloat(r.pctAbove)<=5?'c-green':parseFloat(r.pctAbove)<=10?'c-amber':'c-muted'}}">${{r.pctAbove}}%</td>
    <td class="${{parseFloat(r.upside)>5?'c-green':parseFloat(r.upside)>0?'c-amber':'c-red'}}">${{r.upside}}%</td>
    <td class="c-red">₹${{r.stopLoss}}</td>
    <td>${{r.hvBullish?'<span class="c-green">▲ Bull</span>':'<span class="c-red">▼ Bear</span>'}}</td>
    <td><a href="https://www.tradingview.com/chart/?symbol=NSE:${{r.symbol}}" target="_blank" class="tv-link">Chart ↗</a></td>
  </tr>`).join('');}}

function mkBearRows(arr){{return arr.map(r=>`
  <tr>
    <td><div class="sym">${{r.symbol}}</div><span class="tag tag-${{r.grade==='short'?'short':r.grade==='watch'?'watch-bear':'weak'}}">${{r.grade==='short'?'SHORT':r.grade==='watch'?'WATCH':'WEAK'}}</span></td>
    <td><span style="font-weight:700;font-size:15px;color:${{r.score>=65?'#ef4444':r.score>=45?'#fb923c':'#64748b'}}">${{r.score}}</span><span class="c-muted">/100</span></td>
    <td class="c-muted" style="font-size:11px">${{r.hvDate}}</td>
    <td class="${{r.daysSince<=7?'c-red':r.daysSince<=30?'c-amber':'c-muted'}}">${{r.daysSince}}d</td>
    <td class="c-muted">₹${{r.hvLow}}</td>
    <td style="font-weight:600;color:#f1f5f9">₹${{r.hvHigh}}</td>
    <td><div style="font-weight:700;font-size:14px;color:#f87171">₹${{r.currentPrice}}</div>
        <div class="price-src">${{r.isLive?'<span class="live-dot"></span>live':'<span class="cached-dot"></span>'+r.priceDate}}${{r.volume>0?' · Vol:'+r.volume.toLocaleString():''}}</div></td>
    <td class="${{parseFloat(r.pctBelow)<=5?'c-red':parseFloat(r.pctBelow)<=10?'c-amber':'c-muted'}}">${{r.pctBelow}}%</td>
    <td class="${{parseFloat(r.downside)>5?'c-red':parseFloat(r.downside)>0?'c-amber':'c-purple'}}">${{r.downside}}%</td>
    <td class="c-green">₹${{r.stopLoss}}</td>
    <td>${{r.hvBearish?'<span class="c-red">▼ Bear</span>':'<span class="c-green">▲ Bull</span>'}}</td>
    <td><a href="https://www.tradingview.com/chart/?symbol=NSE:${{r.symbol}}" target="_blank" class="tv-link">Chart ↗</a></td>
  </tr>`).join('');}}

const BULL_HDR=`<thead><tr><th>Symbol</th><th>Score</th><th>HV Date</th><th>Days</th><th>HV Low</th><th>HV High</th><th>Price</th><th>% Above</th><th>Upside</th><th>Stop Loss</th><th>Candle</th><th>Chart</th></tr></thead>`;
const BEAR_HDR=`<thead><tr><th>Symbol</th><th>Score</th><th>HV Date</th><th>Days</th><th>HV Low</th><th>HV High (Res.)</th><th>Price</th><th>% Below</th><th>Downside</th><th>Stop Loss</th><th>Candle</th><th>Chart</th></tr></thead>`;

function renderTable(mode){{
  if(mode==='bull'){{
    const el=document.getElementById('b-results');
    let data=bullResults;
    if(bullView==='buy')data=bullResults.filter(r=>r.grade==='buy');
    if(bullView==='watch')data=bullResults.filter(r=>r.grade==='watch');
    if(bullView==='broken')data=bullResults.filter(r=>r.grade==='broken');
    if(!data.length){{el.innerHTML='<div class="empty"><div style="font-size:16px;font-weight:600;color:#64748b">No stocks match — try relaxing filters</div></div>';return;}}
    if(bullView==='all'){{
      const b=data.filter(r=>r.grade==='buy'),w=data.filter(r=>r.grade==='watch'),x=data.filter(r=>r.grade==='broken');
      let html='';
      if(b.length)html+=`<div class="section-title" style="color:#22c55e">▲ Buy Now <span class="badge">${{b.length}}</span></div><table>${{BULL_HDR}}<tbody>${{mkBullRows(b)}}</tbody></table>`;
      if(w.length)html+=`<div class="section-title" style="color:#f59e0b">◆ Watchlist <span class="badge">${{w.length}}</span></div><table>${{BULL_HDR}}<tbody>${{mkBullRows(w)}}</tbody></table>`;
      if(x.length)html+=`<div class="section-title" style="color:#ef4444">▼ Support Broken <span class="badge">${{x.length}}</span></div><table>${{BULL_HDR}}<tbody>${{mkBullRows(x)}}</tbody></table>`;
      el.innerHTML=html;
    }}else{{el.innerHTML=`<table>${{BULL_HDR}}<tbody>${{mkBullRows(data)}}</tbody></table>`;}}
  }}else{{
    const el=document.getElementById('r-results');
    let data=bearResults;
    if(bearView==='short')data=bearResults.filter(r=>r.grade==='short');
    if(bearView==='watch')data=bearResults.filter(r=>r.grade==='watch');
    if(bearView==='weak')data=bearResults.filter(r=>r.grade==='weak');
    if(!data.length){{el.innerHTML='<div class="empty"><div style="font-size:16px;font-weight:600;color:#64748b">No stocks match — try relaxing filters</div></div>';return;}}
    if(bearView==='all'){{
      const s=data.filter(r=>r.grade==='short'),w=data.filter(r=>r.grade==='watch'),k=data.filter(r=>r.grade==='weak');
      let html='';
      if(s.length)html+=`<div class="section-title" style="color:#ef4444">▼ Short Now <span class="badge">${{s.length}}</span></div><table>${{BEAR_HDR}}<tbody>${{mkBearRows(s)}}</tbody></table>`;
      if(w.length)html+=`<div class="section-title" style="color:#fb923c">◆ Watch <span class="badge">${{w.length}}</span></div><table>${{BEAR_HDR}}<tbody>${{mkBearRows(w)}}</tbody></table>`;
      if(k.length)html+=`<div class="section-title" style="color:#a78bfa">◇ Weak <span class="badge">${{k.length}}</span></div><table>${{BEAR_HDR}}<tbody>${{mkBearRows(k)}}</tbody></table>`;
      el.innerHTML=html;
    }}else{{el.innerHTML=`<table>${{BEAR_HDR}}<tbody>${{mkBearRows(data)}}</tbody></table>`;}}
  }}
}}

function exportCSV(){{
  const isBull=activeTab==='bull',results=isBull?bullResults:bearResults;
  if(!results.length){{alert('Run scan first');return;}}
  const hdr=isBull?['Symbol','Score','Grade','HV Date','Days','HV Low','HV High','Price','Price Date','% Above','Upside%','Stop Loss','Candle','Source']:
                   ['Symbol','Score','Grade','HV Date','Days','HV Low','HV High','Price','Price Date','% Below','Downside%','Stop Loss','Candle','Source'];
  const rows=results.map(r=>isBull?[r.symbol,r.score,r.grade,r.hvDate,r.daysSince,r.hvLow,r.hvHigh,r.currentPrice,r.priceDate,r.pctAbove,r.upside,r.stopLoss,r.hvBullish?'Bullish':'Bearish',r.isLive?'Live':'Stored']:
                                   [r.symbol,r.score,r.grade,r.hvDate,r.daysSince,r.hvLow,r.hvHigh,r.currentPrice,r.priceDate,r.pctBelow,r.downside,r.stopLoss,r.hvBearish?'Bearish':'Bullish',r.isLive?'Live':'Stored']);
  const csv=[hdr,...rows].map(r=>r.join(',')).join('\\n');
  const a=document.createElement('a');
  a.href='data:text/csv;charset=utf-8,'+encodeURIComponent(csv);
  a.download=`NSE_${{isBull?'Bull':'Bear'}}_Scan_${{new Date().toISOString().slice(0,10)}}.csv`;
  a.click();
}}
</script>
</body>
</html>'''
    return html

def save_html(html):
    os.makedirs(SERVE_DIR, exist_ok=True)
    out = os.path.join(SERVE_DIR, 'scanner_data.html')
    with open(out, 'w', encoding='utf-8') as f:
        f.write(html)
    also = OUTPUT_HTML
    with open(also, 'w', encoding='utf-8') as f:
        f.write(html)
    log(f"Saved: {out}")
    log(f"Saved: {also}")
    return out

def main():
    log("=" * 50)
    log("NSE Volume Scanner — HTML Builder")
    log("=" * 50)
    try:
        excel = find_excel()
        raw   = build_raw_data(excel)
        html  = generate_html(raw)
        out   = save_html(html)
        log(f"Done! {len(raw)} stocks embedded.")
        log(f"Open: {out}")
    except Exception as e:
        log(f"ERROR: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()
