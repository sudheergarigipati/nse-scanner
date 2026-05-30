#!/usr/bin/env python3
# ═══════════════════════════════════════════════════════════════
#  db_manager.py
#  Manages SQLite database of daily OHLCV data for all NSE stocks
#  - First run: imports 5yr history from Excel + backfills via yfinance
#  - Daily run: fetches today's data and updates database
#  - Recalculates highest volume day per stock automatically
#
#  Fix: NIFTY50 and BANKNIFTY are indices — not stocks
#  They use ^NSEI and ^NSEBANK on yfinance, not .NS suffix
# ═══════════════════════════════════════════════════════════════

import sqlite3
import yfinance as yf
import pandas as pd
import openpyxl
import os
import sys
import glob
import json
import time
from datetime import datetime, date, timedelta

BASE_DIR = os.path.expanduser('~/nse-scanner')
DB_PATH  = os.path.join(BASE_DIR, 'nse_data.db')
LOG_FILE = os.path.join(BASE_DIR, 'logs/db_manager.log')

TELEGRAM_TOKEN   = '8788684553:AAHfZ_q0Hh2mdUNOwELu_PQPePpptKtixGM'
TELEGRAM_CHAT_ID = '-5282064943'

# ── Index symbols — these use different yfinance symbols ──────
# Never append .NS to these
INDEX_SYMBOLS = {
    'NIFTY50':   '^NSEI',
    'BANKNIFTY': '^NSEBANK',
}

def get_yf_symbol(sym):
    """Return the correct yfinance symbol for a given NSE symbol."""
    return INDEX_SYMBOLS.get(sym, f"{sym}.NS")

def is_index(sym):
    return sym in INDEX_SYMBOLS

def send_telegram(message):
    try:
        import urllib.request, urllib.parse, json as _json
        url  = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
        data = urllib.parse.urlencode({
            'chat_id': TELEGRAM_CHAT_ID,
            'text':    message,
        }).encode()
        req = urllib.request.Request(url, data=data, method='POST')
        with urllib.request.urlopen(req, timeout=10) as r:
            result = _json.loads(r.read())
            return result.get('ok', False)
    except Exception as e:
        log(f"Telegram error: {e}")
        return False

def log(msg):
    ts   = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f"[{ts}] {msg}"
    print(line)
    os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)
    with open(LOG_FILE, 'a') as f:
        f.write(line + '\n')

# ══════════════════════════════════════════════════════════════
#  DATABASE SETUP
# ══════════════════════════════════════════════════════════════
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c    = conn.cursor()

    c.execute('''CREATE TABLE IF NOT EXISTS daily_prices (
        symbol      TEXT NOT NULL,
        date        TEXT NOT NULL,
        open        REAL,
        high        REAL,
        low         REAL,
        close       REAL,
        volume      INTEGER,
        PRIMARY KEY (symbol, date)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS hv_summary (
        symbol          TEXT PRIMARY KEY,
        hv_date         TEXT,
        hv_volume       INTEGER,
        hv_open         REAL,
        hv_high         REAL,
        hv_low          REAL,
        hv_close        REAL,
        latest_date     TEXT,
        latest_close    REAL,
        latest_volume   INTEGER,
        total_days      INTEGER,
        updated_at      TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS hv_alerts (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        symbol      TEXT,
        alert_date  TEXT,
        new_volume  INTEGER,
        old_volume  INTEGER,
        old_hv_date TEXT,
        price       REAL,
        alerted_at  TEXT
    )''')

    c.execute('CREATE INDEX IF NOT EXISTS idx_daily_symbol ON daily_prices(symbol)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_daily_date   ON daily_prices(date)')
    conn.commit()
    conn.close()
    log("Database initialized")

# ══════════════════════════════════════════════════════════════
#  STOCK LIST
# ══════════════════════════════════════════════════════════════
def get_stock_list():
    """Get list of stocks from DB. Index symbols handled separately."""
    conn = get_db()
    c    = conn.cursor()
    c.execute('SELECT DISTINCT symbol FROM hv_summary ORDER BY symbol')
    syms = [r[0] for r in c.fetchall()]
    conn.close()

    if syms:
        return syms

    # Fall back to Excel
    files = glob.glob(os.path.join(BASE_DIR, 'NSE_5yr*.xlsx')) + \
            glob.glob(os.path.join(BASE_DIR, '*.xlsx'))
    if not files:
        log("ERROR: No Excel file and no stocks in DB")
        return []

    excel   = max(files, key=os.path.getmtime)
    wb      = openpyxl.load_workbook(excel, read_only=True)
    ws_name = next((n for n in wb.sheetnames if 'SUMMARY' in n.upper()), None)
    if not ws_name:
        return []
    ws   = wb[ws_name]
    rows = list(ws.iter_rows(values_only=True))
    syms = [str(r[0]).strip() for r in rows[1:] if r[0]]
    log(f"Got {len(syms)} stocks from Excel")
    return syms

# ══════════════════════════════════════════════════════════════
#  IMPORT FROM EXCEL (one-time)
# ══════════════════════════════════════════════════════════════
def import_from_excel():
    files = glob.glob(os.path.join(BASE_DIR, 'NSE_5yr*.xlsx')) + \
            glob.glob(os.path.join(BASE_DIR, '*.xlsx'))
    if not files:
        log("No Excel file found — skipping Excel import")
        return 0

    excel   = max(files, key=os.path.getmtime)
    log(f"Importing from Excel: {os.path.basename(excel)}")

    wb      = openpyxl.load_workbook(excel, read_only=True)
    ws_name = next((n for n in wb.sheetnames if 'SUMMARY' in n.upper()), None)
    if not ws_name:
        log("No SUMMARY sheet found")
        return 0

    ws    = wb[ws_name]
    rows  = list(ws.iter_rows(values_only=True))
    conn  = get_db()
    c     = conn.cursor()
    count = 0

    for row in rows[1:]:
        try:
            sym, fd, td, days, mv, mvd, hl, hh, hc, lc, ld = row
            if not sym or not mv:
                continue
            c.execute('''INSERT OR IGNORE INTO hv_summary
                (symbol, hv_date, hv_volume, hv_low, hv_high, hv_close,
                 latest_date, latest_close, total_days, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?)''',
                (str(sym).strip(), str(mvd).strip(), int(mv),
                 float(hl), float(hh), float(hc),
                 str(ld).strip(), float(lc),
                 int(days) if days else 0,
                 datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            count += 1
        except Exception as e:
            continue

    conn.commit()
    conn.close()
    log(f"Imported {count} stocks from Excel into DB")
    return count

# ══════════════════════════════════════════════════════════════
#  BACKFILL HISTORY FROM YFINANCE (one-time)
# ══════════════════════════════════════════════════════════════
def backfill_history(symbols, years=5):
    log(f"Starting backfill for {len(symbols)} stocks ({years} years)...")
    conn  = get_db()
    c     = conn.cursor()
    done  = 0
    BATCH = 50

    start = (date.today() - timedelta(days=years*365)).strftime("%Y-%m-%d")
    end   = (date.today() + timedelta(days=1)).strftime("%Y-%m-%d")

    # Separate indices from regular stocks
    index_syms = [s for s in symbols if is_index(s)]
    stock_syms = [s for s in symbols if not is_index(s)]

    # ── Backfill index symbols one by one ─────────────────────
    for sym in index_syms:
        yf_sym = get_yf_symbol(sym)
        log(f"  Backfilling index {sym} ({yf_sym})...")
        try:
            ticker = yf.Ticker(yf_sym)
            df     = ticker.history(start=start, end=end,
                                    interval='1d', auto_adjust=True)
            if df is None or df.empty:
                log(f"  {sym}: no data")
                continue
            rows_to_insert = []
            for idx, row in df.iterrows():
                if pd.isna(row['Close']) or row['Close'] <= 0:
                    continue
                rows_to_insert.append((
                    sym,
                    str(idx.date()),
                    round(float(row['Open']),  2) if not pd.isna(row['Open'])   else None,
                    round(float(row['High']),  2) if not pd.isna(row['High'])   else None,
                    round(float(row['Low']),   2) if not pd.isna(row['Low'])    else None,
                    round(float(row['Close']), 2),
                    int(row['Volume'])            if not pd.isna(row['Volume']) else 0,
                ))
            c.executemany('''INSERT OR REPLACE INTO daily_prices
                (symbol, date, open, high, low, close, volume)
                VALUES (?,?,?,?,?,?,?)''', rows_to_insert)
            conn.commit()
            log(f"  {sym}: {len(rows_to_insert)} bars saved")
            done += 1
        except Exception as e:
            log(f"  {sym} backfill error: {e}")

    # ── Backfill regular stocks in batches ────────────────────
    for i in range(0, len(stock_syms), BATCH):
        batch       = stock_syms[i:i+BATCH]
        tickers_str = ' '.join([f"{s}.NS" for s in batch])

        try:
            data = yf.download(
                tickers_str,
                start=start,
                end=end,
                interval='1d',
                group_by='ticker',
                auto_adjust=True,
                progress=False,
                threads=True
            )

            for sym in batch:
                try:
                    ticker_key = f"{sym}.NS"
                    df = data[ticker_key] if len(batch) > 1 else data
                    if df is None or df.empty:
                        continue

                    rows_to_insert = []
                    for idx, row in df.iterrows():
                        if pd.isna(row['Close']) or row['Close'] <= 0:
                            continue
                        rows_to_insert.append((
                            sym,
                            str(idx.date()),
                            round(float(row['Open']),  2) if not pd.isna(row['Open'])   else None,
                            round(float(row['High']),  2) if not pd.isna(row['High'])   else None,
                            round(float(row['Low']),   2) if not pd.isna(row['Low'])    else None,
                            round(float(row['Close']), 2),
                            int(row['Volume'])            if not pd.isna(row['Volume']) else 0,
                        ))
                    c.executemany('''INSERT OR REPLACE INTO daily_prices
                        (symbol, date, open, high, low, close, volume)
                        VALUES (?,?,?,?,?,?,?)''', rows_to_insert)
                    done += 1
                except Exception:
                    continue

        except Exception as e:
            log(f"Batch error: {e}")
            continue

        conn.commit()
        log(f"  Backfill progress: {min(i+BATCH, len(stock_syms))}/{len(stock_syms)} stocks")
        time.sleep(0.3)

    conn.close()
    log(f"Backfill complete: {done} stocks with historical data")
    return done

# ══════════════════════════════════════════════════════════════
#  DAILY UPDATE
# ══════════════════════════════════════════════════════════════
def fetch_daily_update(symbols):
    log(f"Fetching daily update for {len(symbols)} stocks...")
    conn  = get_db()
    c     = conn.cursor()
    done  = 0
    BATCH = 50

    # Get last known date
    conn2 = get_db()
    c2    = conn2.cursor()
    c2.execute("SELECT MAX(date) FROM daily_prices")
    last_date = c2.fetchone()[0]
    conn2.close()

    if last_date:
        last_dt     = datetime.strptime(last_date, "%Y-%m-%d").date()
        days_missed = (date.today() - last_dt).days
        start       = last_dt.strftime("%Y-%m-%d")
        log(f"Last data: {last_date} | Days to fetch: {days_missed}")
    else:
        start = (date.today() - timedelta(days=365)).strftime("%Y-%m-%d")
        log("No existing data — fetching 1 year")

    end = (date.today() + timedelta(days=1)).strftime("%Y-%m-%d")

    # Separate indices from regular stocks
    index_syms = [s for s in symbols if is_index(s)]
    stock_syms = [s for s in symbols if not is_index(s)]

    # ── Update index symbols one by one ───────────────────────
    for sym in index_syms:
        yf_sym = get_yf_symbol(sym)
        log(f"  Updating index {sym} ({yf_sym})...")
        try:
            ticker = yf.Ticker(yf_sym)
            df     = ticker.history(start=start, end=end,
                                    interval='1d', auto_adjust=True)
            if df is None or df.empty:
                log(f"  {sym}: no data returned")
                continue
            rows_to_insert = []
            for idx, row in df.iterrows():
                if pd.isna(row['Close']) or row['Close'] <= 0:
                    continue
                rows_to_insert.append((
                    sym,
                    str(idx.date()),
                    round(float(row['Open']),  2) if not pd.isna(row['Open'])   else None,
                    round(float(row['High']),  2) if not pd.isna(row['High'])   else None,
                    round(float(row['Low']),   2) if not pd.isna(row['Low'])    else None,
                    round(float(row['Close']), 2),
                    int(row['Volume'])            if not pd.isna(row['Volume']) else 0,
                ))
            if rows_to_insert:
                c.executemany('''INSERT OR REPLACE INTO daily_prices
                    (symbol, date, open, high, low, close, volume)
                    VALUES (?,?,?,?,?,?,?)''', rows_to_insert)
                conn.commit()
                log(f"  {sym}: {len(rows_to_insert)} bars updated")
            done += 1
        except Exception as e:
            log(f"  {sym} update error: {e}")

    # ── Update regular stocks in batches ──────────────────────
    for i in range(0, len(stock_syms), BATCH):
        batch       = stock_syms[i:i+BATCH]
        tickers_str = ' '.join([f"{s}.NS" for s in batch])

        try:
            data = yf.download(
                tickers_str,
                start=start,
                end=end,
                interval='1d',
                group_by='ticker',
                auto_adjust=True,
                progress=False,
                threads=True
            )

            for sym in batch:
                try:
                    ticker_key = f"{sym}.NS"
                    df = data[ticker_key] if len(batch) > 1 else data
                    if df is None or df.empty:
                        continue

                    rows_to_insert = []
                    for idx, row in df.iterrows():
                        if pd.isna(row['Close']) or row['Close'] <= 0:
                            continue
                        rows_to_insert.append((
                            sym,
                            str(idx.date()),
                            round(float(row['Open']),  2) if not pd.isna(row['Open'])   else None,
                            round(float(row['High']),  2) if not pd.isna(row['High'])   else None,
                            round(float(row['Low']),   2) if not pd.isna(row['Low'])    else None,
                            round(float(row['Close']), 2),
                            int(row['Volume'])            if not pd.isna(row['Volume']) else 0,
                        ))

                    c.executemany('''INSERT OR REPLACE INTO daily_prices
                        (symbol, date, open, high, low, close, volume)
                        VALUES (?,?,?,?,?,?,?)''', rows_to_insert)
                    done += 1

                except Exception:
                    continue

        except Exception as e:
            log(f"Batch error: {e}")
            continue

        conn.commit()
        time.sleep(0.3)

    conn.close()
    log(f"Daily update complete: {done}/{len(symbols)} stocks updated")
    return done

# ══════════════════════════════════════════════════════════════
#  RECALCULATE HV SUMMARY
# ══════════════════════════════════════════════════════════════
def recalculate_hv_summary():
    log("Recalculating HV summary...")
    conn = get_db()
    c    = conn.cursor()

    c.execute('SELECT DISTINCT symbol FROM daily_prices')
    symbols = [r[0] for r in c.fetchall()]

    new_hv_alerts = []
    updated       = 0

    for sym in symbols:
        # Skip indices — no HV concept for Nifty/BankNifty
        if is_index(sym):
            continue
        try:
            c.execute('SELECT hv_date, hv_volume FROM hv_summary WHERE symbol=?', (sym,))
            old         = c.fetchone()
            old_hv_vol  = old['hv_volume'] if old else 0
            old_hv_date = old['hv_date']   if old else None

            five_yr_ago = (date.today() - timedelta(days=5*365)).strftime('%Y-%m-%d')
            c.execute('''SELECT date, open, high, low, close, volume
                         FROM daily_prices
                         WHERE symbol=? AND date >= ? AND volume > 0
                         ORDER BY volume DESC LIMIT 1''', (sym, five_yr_ago))
            hv = c.fetchone()
            if not hv:
                continue

            c.execute('''SELECT date, close, volume FROM daily_prices
                         WHERE symbol=? ORDER BY date DESC LIMIT 1''', (sym,))
            latest = c.fetchone()
            if not latest:
                continue

            c.execute('SELECT COUNT(*) FROM daily_prices WHERE symbol=? AND date >= ?',
                      (sym, five_yr_ago))
            total_days = c.fetchone()[0]

            if hv['volume'] > old_hv_vol and old_hv_date and hv['date'] != old_hv_date:
                new_hv_alerts.append({
                    'symbol':      sym,
                    'alert_date':  str(date.today()),
                    'new_volume':  hv['volume'],
                    'old_volume':  old_hv_vol,
                    'old_hv_date': old_hv_date,
                    'price':       latest['close'],
                })
                log(f"  NEW HV: {sym} | {hv['volume']:,} > {old_hv_vol:,} (old: {old_hv_date})")

            c.execute('''INSERT OR REPLACE INTO hv_summary
                (symbol, hv_date, hv_volume, hv_high, hv_low, hv_close,
                 latest_date, latest_close, latest_volume, total_days, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                (sym, hv['date'], hv['volume'],
                 hv['high'], hv['low'], hv['close'],
                 latest['date'], latest['close'], latest['volume'],
                 total_days,
                 datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            updated += 1

        except Exception:
            continue

    for alert in new_hv_alerts:
        c.execute('''INSERT INTO hv_alerts
            (symbol, alert_date, new_volume, old_volume, old_hv_date, price, alerted_at)
            VALUES (?,?,?,?,?,?,?)''',
            (alert['symbol'], alert['alert_date'], alert['new_volume'],
             alert['old_volume'], alert['old_hv_date'], alert['price'],
             datetime.now().strftime('%Y-%m-%d %H:%M:%S')))

    conn.commit()
    conn.close()
    log(f"HV summary updated: {updated} stocks | {len(new_hv_alerts)} new HV alerts")

    for alert in new_hv_alerts:
        ist_time = (datetime.now()).strftime('%d %b %Y %H:%M IST')
        msg = (
            f"NEW 5-YEAR HIGHEST VOLUME!\n"
            f"{alert['symbol']}\n"
            f"{ist_time}\n\n"
            f"Today Vol   : {alert['new_volume']:,}\n"
            f"Previous HV : {alert['old_volume']:,} ({alert['old_hv_date']})\n"
            f"Price       : Rs {alert['price']}\n\n"
            f"Chart: https://www.tradingview.com/chart/?symbol=NSE:{alert['symbol']}"
        )
        send_telegram(msg)
        log(f"  Telegram sent for {alert['symbol']}")

    return new_hv_alerts

# ══════════════════════════════════════════════════════════════
#  CHECK TODAY'S NEW HV (intraday)
# ══════════════════════════════════════════════════════════════
def check_intraday_new_hv(symbols):
    log("Checking intraday new HV...")
    conn   = get_db()
    c      = conn.cursor()
    alerts = []
    BATCH  = 50

    # Only check regular stocks — not indices
    stock_syms = [s for s in symbols if not is_index(s)]

    for i in range(0, len(stock_syms), BATCH):
        batch       = stock_syms[i:i+BATCH]
        tickers_str = ' '.join([f"{s}.NS" for s in batch])

        try:
            data = yf.download(
                tickers_str,
                period='1d',
                interval='1d',
                group_by='ticker',
                auto_adjust=True,
                progress=False,
                threads=True
            )

            for sym in batch:
                try:
                    ticker_key = f"{sym}.NS"
                    df = data[ticker_key] if len(batch) > 1 else data
                    if df is None or df.empty:
                        continue

                    today_vol   = int(df.iloc[-1]['Volume'])
                    today_close = float(df.iloc[-1]['Close'])
                    if today_vol <= 0:
                        continue

                    c.execute('SELECT hv_date, hv_volume FROM hv_summary WHERE symbol=?', (sym,))
                    row = c.fetchone()
                    if not row:
                        continue

                    hv_vol  = row['hv_volume']
                    hv_date = row['hv_date']

                    now          = datetime.now()
                    market_open  = now.replace(hour=9,  minute=15, second=0)
                    market_close = now.replace(hour=15, minute=30, second=0)
                    elapsed      = max((now - market_open).total_seconds(), 1)
                    total        = (market_close - market_open).total_seconds()
                    pct_day      = min(elapsed / total, 1.0)
                    projected    = int(today_vol / pct_day) if pct_day > 0 else today_vol

                    if today_vol > hv_vol:
                        alerts.append({
                            'symbol':    sym,
                            'status':    'CONFIRMED',
                            'today_vol': today_vol,
                            'hv_vol':    hv_vol,
                            'hv_date':   hv_date,
                            'price':     today_close,
                            'pct_above': round((today_vol - hv_vol) / hv_vol * 100, 1),
                            'projected': today_vol,
                        })
                    elif projected > hv_vol and pct_day < 0.95:
                        alerts.append({
                            'symbol':    sym,
                            'status':    'ON_TRACK',
                            'today_vol': today_vol,
                            'hv_vol':    hv_vol,
                            'hv_date':   hv_date,
                            'price':     today_close,
                            'pct_above': round((projected - hv_vol) / hv_vol * 100, 1),
                            'projected': projected,
                        })

                except Exception:
                    continue

        except Exception as e:
            log(f"Intraday batch error: {e}")

        time.sleep(0.3)

    conn.close()
    alerts.sort(key=lambda x: (x['status'] == 'CONFIRMED', x['today_vol']), reverse=True)
    log(f"Intraday HV check: {len(alerts)} alerts "
        f"({sum(1 for a in alerts if a['status']=='CONFIRMED')} confirmed)")
    return alerts

# ══════════════════════════════════════════════════════════════
#  EXPORT SUMMARY JSON (for scanner HTML)
# ══════════════════════════════════════════════════════════════
def export_summary_json():
    conn = get_db()
    c    = conn.cursor()
    c.execute('''SELECT symbol, hv_date, hv_volume, hv_high, hv_low, hv_close,
                        latest_date, latest_close, latest_volume, total_days
                 FROM hv_summary
                 WHERE hv_date IS NOT NULL AND latest_close > 0''')
    rows = c.fetchall()
    conn.close()

    raw = []
    for r in rows:
        raw.append({
            's':   r['symbol'],
            'mvd': r['hv_date'],
            'mv':  r['hv_volume']  or 0,
            'hh':  r['hv_high']   or 0,
            'hl':  r['hv_low']    or 0,
            'hc':  r['hv_close']  or 0,
            'lc':  r['latest_close']  or 0,
            'ld':  r['latest_date']   or '',
            'lv':  r['latest_volume'] or 0,
            'td':  r['total_days']    or 0,
        })

    out_path = os.path.join(BASE_DIR, 'summary.json')
    with open(out_path, 'w') as f:
        json.dump(raw, f, separators=(',', ':'))

    log(f"Exported {len(raw)} stocks to summary.json")
    return raw

# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════
def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--init',     action='store_true', help='Initialize DB and import Excel')
    parser.add_argument('--backfill', action='store_true', help='Backfill 5yr history from yfinance')
    parser.add_argument('--update',   action='store_true', help='Fetch today data + recalculate HV')
    parser.add_argument('--intraday', action='store_true', help='Check intraday new HV alerts')
    parser.add_argument('--export',   action='store_true', help='Export summary JSON')
    args = parser.parse_args()

    os.makedirs(BASE_DIR, exist_ok=True)
    os.makedirs(os.path.join(BASE_DIR, 'logs'), exist_ok=True)

    log("=" * 55)
    log(f"DB Manager — {datetime.now().strftime('%d %b %Y %H:%M')}")
    log("=" * 55)

    if args.init:
        init_db()
        count = import_from_excel()
        if count == 0:
            log("No Excel data — run --backfill to fetch from yfinance")
        else:
            log(f"Init complete: {count} stocks imported")

    elif args.backfill:
        init_db()
        symbols = get_stock_list()
        # Always include indices in backfill
        for idx in INDEX_SYMBOLS:
            if idx not in symbols:
                symbols.append(idx)
        if not symbols:
            log("No stocks to backfill")
            return
        done = backfill_history(symbols)
        recalculate_hv_summary()
        export_summary_json()
        log(f"Backfill complete: {done} stocks")

    elif args.update:
        init_db()
        symbols = get_stock_list()
        # Always include indices in daily update
        for idx in INDEX_SYMBOLS:
            if idx not in symbols:
                symbols.append(idx)
        if not symbols:
            log("No stocks to update")
            return
        fetch_daily_update(symbols)
        recalculate_hv_summary()
        export_summary_json()

    elif args.intraday:
        symbols = get_stock_list()
        alerts  = check_intraday_new_hv(symbols)
        log(f"Intraday alerts: {len(alerts)}")

    elif args.export:
        export_summary_json()

    else:
        log("Usage: db_manager.py --init | --backfill | --update | --intraday | --export")


if __name__ == '__main__':
    main()
